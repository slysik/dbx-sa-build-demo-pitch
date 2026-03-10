# /// script
# dependencies = ["openpyxl"]
# ///
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colors ──
HEADER_FILL = PatternFill('solid', fgColor='1B3A5C')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
SUBHEADER_FILL = PatternFill('solid', fgColor='2C5F8A')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
GA_FILL = PatternFill('solid', fgColor='C6EFCE')
GA_FONT = Font(color='006100', size=10, name='Calibri')
PREVIEW_FILL = PatternFill('solid', fgColor='FFEB9C')
PREVIEW_FONT = Font(color='9C6500', size=10, name='Calibri')
BETA_FILL = PatternFill('solid', fgColor='FFC7CE')
BETA_FONT = Font(color='9C0006', size=10, name='Calibri')
NA_FILL = PatternFill('solid', fgColor='D9D9D9')
NA_FONT = Font(color='808080', size=10, name='Calibri')
BODY_FONT = Font(size=10, name='Calibri')
BOLD_FONT = Font(bold=True, size=10, name='Calibri')
SECTION_FILL = PatternFill('solid', fgColor='E8EFF5')
SECTION_FONT = Font(bold=True, size=10, name='Calibri', color='1B3A5C')
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_status_cell(cell, status):
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if status == 'GA':
        cell.fill = GA_FILL
        cell.font = GA_FONT
    elif status == 'Preview':
        cell.fill = PREVIEW_FILL
        cell.font = PREVIEW_FONT
    elif status == 'Beta':
        cell.fill = BETA_FILL
        cell.font = BETA_FONT
    elif status == '—':
        cell.fill = NA_FILL
        cell.font = NA_FONT
    else:
        cell.font = BODY_FONT

# ════════════════════════════════════════════════════
# Sheet 1: Cloud Availability Matrix
# ════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Cloud Availability'

headers = ['Connector', 'Category', 'AWS', 'Azure', 'GCP', 'Notes']
connectors = [
    # SaaS - CRM & Sales
    ['Salesforce', 'CRM', 'GA', 'GA', 'GA', 'Formula fields need special config for incremental'],
    ['Dynamics 365', 'CRM', 'Preview', 'Preview', 'Preview', 'Row filtering supported'],
    ['HubSpot', 'CRM', 'Beta', 'Beta', 'Beta', 'API-only authoring; partial incremental'],
    # SaaS - IT & Ops
    ['ServiceNow', 'IT/Ops', 'GA', 'GA', 'GA', 'Cursor field required for incremental'],
    ['Jira', 'IT/Ops', 'Beta', 'Beta', '—', 'Not available on GCP yet'],
    ['Confluence', 'IT/Ops', 'Beta', 'Beta', 'Beta', ''],
    ['Zendesk Support', 'IT/Ops', 'Beta', 'Beta', 'Beta', ''],
    # SaaS - HR & Finance
    ['Workday Reports', 'HR/Finance', 'GA', 'GA', 'GA', 'Supports basic auth + OAuth refresh token'],
    ['NetSuite', 'Finance', 'Preview', 'Preview', '—', 'Service account JSON key auth; 200 table limit'],
    # SaaS - Marketing & Ads
    ['Google Analytics', 'Marketing', 'GA', 'GA', 'GA', ''],
    ['Google Ads', 'Ads', 'Beta', 'Beta', 'Beta', 'API-only; report tables incremental, resource tables full refresh'],
    ['Meta Ads', 'Ads', 'Beta', 'Beta', 'Beta', 'API-only; no SCD Type 2'],
    ['TikTok Ads', 'Ads', 'Beta', 'Beta', 'Beta', 'API-only authoring'],
    # SaaS - Collaboration
    ['SharePoint', 'Collab', 'Beta', 'Beta', 'Beta', 'OAuth M2M in Preview on Azure'],
    # Databases
    ['SQL Server', 'Database', 'GA', 'GA', 'GA', 'CDC + change tracking; gateway on classic compute'],
    ['MySQL', 'Database', 'Preview', 'Preview', 'Preview', 'Username/password auth only'],
    ['PostgreSQL', 'Database', 'Preview', 'Preview', '—', 'Not available on GCP yet'],
]

# Title
ws1.merge_cells('A1:F1')
title_cell = ws1['A1']
title_cell.value = 'Lakeflow Connect — Managed Connector Availability by Cloud'
title_cell.font = Font(bold=True, size=14, name='Calibri', color='1B3A5C')
title_cell.alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('A2:F2')
ws1['A2'].value = 'As of February 2026 | GA = Generally Available | Preview = Public Preview | Beta = Beta | — = Not Available'
ws1['A2'].font = Font(size=9, name='Calibri', color='666666', italic=True)

# Headers row 4
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# Data
row = 5
current_category = None
for conn in connectors:
    name, category, aws, azure, gcp, notes = conn

    ws1.cell(row=row, column=1, value=name).font = BOLD_FONT
    ws1.cell(row=row, column=1).alignment = LEFT
    ws1.cell(row=row, column=1).border = THIN_BORDER

    ws1.cell(row=row, column=2, value=category).font = BODY_FONT
    ws1.cell(row=row, column=2).alignment = CENTER
    ws1.cell(row=row, column=2).border = THIN_BORDER

    style_status_cell(ws1.cell(row=row, column=3, value=aws), aws)
    style_status_cell(ws1.cell(row=row, column=4, value=azure), azure)
    style_status_cell(ws1.cell(row=row, column=5, value=gcp), gcp)

    ws1.cell(row=row, column=6, value=notes).font = Font(size=9, name='Calibri', color='555555')
    ws1.cell(row=row, column=6).alignment = LEFT
    ws1.cell(row=row, column=6).border = THIN_BORDER
    row += 1

# Summary row
row += 1
ws1.cell(row=row, column=1, value='TOTALS').font = Font(bold=True, size=11, name='Calibri')
for col, cloud in [(3, 'AWS'), (4, 'Azure'), (5, 'GCP')]:
    count = sum(1 for c in connectors if c[col-1] != '—')
    cell = ws1.cell(row=row, column=col, value=count)
    cell.font = Font(bold=True, size=11, name='Calibri')
    cell.alignment = CENTER
    cell.border = THIN_BORDER

row += 1
ws1.cell(row=row, column=1, value='GA count').font = BODY_FONT
for col in [3, 4, 5]:
    count = sum(1 for c in connectors if c[col-1] == 'GA')
    cell = ws1.cell(row=row, column=col, value=count)
    cell.font = GA_FONT
    cell.fill = GA_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

row += 1
ws1.cell(row=row, column=1, value='Preview count').font = BODY_FONT
for col in [3, 4, 5]:
    count = sum(1 for c in connectors if c[col-1] == 'Preview')
    cell = ws1.cell(row=row, column=col, value=count)
    cell.font = PREVIEW_FONT
    cell.fill = PREVIEW_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

row += 1
ws1.cell(row=row, column=1, value='Beta count').font = BODY_FONT
for col in [3, 4, 5]:
    count = sum(1 for c in connectors if c[col-1] == 'Beta')
    cell = ws1.cell(row=row, column=col, value=count)
    cell.font = BETA_FONT
    cell.fill = BETA_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 55

# ════════════════════════════════════════════════════
# Sheet 2: Feature Comparison
# ════════════════════════════════════════════════════
ws2 = wb.create_sheet('Feature Matrix')

features = [
    'UI Authoring', 'API Authoring', 'Asset Bundles', 'Incremental Ingestion',
    'UC Governance', 'Workflows Orch.', 'SCD Type 2', 'Column Select',
    'Row Filtering', 'Schema Evo: New Cols', 'Schema Evo: Type Changes',
    'Schema Evo: Renames', 'Schema Evo: New Tables', 'Max Tables/Pipeline'
]

feature_data = {
    'Salesforce':      ['Y','Y','Y','Y*','Y','Y','Y','Y','Y','Y','N','Y','N/A','250'],
    'ServiceNow':      ['Y','Y','Y','Y*','Y','Y','Y','Y','Y','Y','N','Y','Y','250'],
    'Google Analytics': ['Y','Y','Y','Y','Y','Y','Y','Y','Y','Y','N','Y','Y','250'],
    'SQL Server':      ['Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N','N','Y','250'],
    'Workday Reports': ['Y','Y','Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N/A','250'],
    'Dynamics 365':    ['Y','Y','Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N/A','250'],
    'MySQL':           ['Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N','Y','Y','250'],
    'PostgreSQL':      ['Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N','Y','Y','250'],
    'NetSuite':        ['Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N','Y','Y','200'],
    'Confluence':      ['Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N','Y','Y','250'],
    'Jira':            ['Y','Y','Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N/A','250'],
    'SharePoint':      ['Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N','N','Y','250'],
    'Zendesk Support': ['Y','Y','Y','Y','Y','Y','Y','Y','Y','Y','N','Y','N/A','250'],
    'HubSpot':         ['N','Y','Y','Partial','Y','Y','N','Y','N','N','N','N','Y','250'],
    'Google Ads':      ['N','Y','Y','Partial','Y','Y','N','Y','N','N','N','N','N','250'],
    'Meta Ads':        ['N','Y','Y','Partial','Y','Y','N','Y','N','N','N','N','N/A','250'],
    'TikTok Ads':      ['N','Y','Y','Y*','Y','Y','Y','Y','N','Y','N','N','Y','250'],
}

ws2.merge_cells('A1:P1')
ws2['A1'].value = 'Lakeflow Connect — Feature Comparison Matrix'
ws2['A1'].font = Font(bold=True, size=14, name='Calibri', color='1B3A5C')

# Headers
ws2.cell(row=3, column=1, value='Connector').fill = HEADER_FILL
ws2.cell(row=3, column=1).font = HEADER_FONT
ws2.cell(row=3, column=1).alignment = CENTER
ws2.cell(row=3, column=1).border = THIN_BORDER

for col, feat in enumerate(features, 2):
    cell = ws2.cell(row=3, column=col, value=feat)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=45)
    cell.border = THIN_BORDER

row = 4
for connector, vals in feature_data.items():
    cell = ws2.cell(row=row, column=1, value=connector)
    cell.font = BOLD_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER

    for col, val in enumerate(vals, 2):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if val == 'Y':
            cell.fill = GA_FILL
            cell.font = GA_FONT
        elif val in ('Y*', 'Partial'):
            cell.fill = PREVIEW_FILL
            cell.font = PREVIEW_FONT
        elif val == 'N':
            cell.fill = BETA_FILL
            cell.font = BETA_FONT
        elif val == 'N/A':
            cell.fill = NA_FILL
            cell.font = NA_FONT
        else:
            cell.font = BODY_FONT
    row += 1

ws2.column_dimensions['A'].width = 22
for c in range(2, len(features)+2):
    ws2.column_dimensions[get_column_letter(c)].width = 14
ws2.row_dimensions[3].height = 70

# ════════════════════════════════════════════════════
# Sheet 3: Standard vs Managed
# ════════════════════════════════════════════════════
ws3 = wb.create_sheet('Standard vs Managed')

ws3.merge_cells('A1:E1')
ws3['A1'].value = 'Lakeflow Connect — Standard Connectors vs Managed Connectors'
ws3['A1'].font = Font(bold=True, size=14, name='Calibri', color='1B3A5C')

headers3 = ['Connector', 'Type', 'Languages', 'Use Case', 'Key Distinction']
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

standard_data = [
    ['Auto Loader (Structured Streaming)', 'Standard', 'Python, Scala', 'Incremental file ingestion from cloud storage', 'cloudFiles format; file notification or directory listing mode'],
    ['Auto Loader (SDP)', 'Standard', 'Python, Scala, SQL', 'File ingestion inside declarative pipelines', 'Same engine, declarative wrapper'],
    ['Auto Loader (DBSQL)', 'Standard', 'SQL', 'File ingestion via SQL warehouses', 'Serverless SQL interface'],
    ['SFTP Ingest', 'Standard', 'Python, SQL', 'Pull files from SFTP servers', 'New addition; file-based sources'],
    ['Kafka (Structured Streaming)', 'Standard', 'Python, Scala', 'Real-time event streaming', 'Low-latency; offset management'],
    ['Kafka (SDP)', 'Standard', 'Python, Scala, SQL', 'Kafka inside declarative pipelines', 'Declarative wrapper over Kafka source'],
    ['Kafka (DBSQL)', 'Standard', 'SQL', 'Kafka via SQL warehouses', 'Serverless SQL interface'],
    ['Amazon Kinesis', 'Standard', 'Python, Scala, SQL', 'AWS-native event streaming', 'AWS only'],
    ['Google Pub/Sub', 'Standard', 'Python, Scala', 'GCP-native event streaming', 'GCP only'],
    ['', '', '', '', ''],
    ['Salesforce', 'Managed', 'No-code / API', 'CRM data ingestion', 'Zero-code UI; fully managed CDC'],
    ['SQL Server', 'Managed', 'No-code / API', 'Database CDC ingestion', 'Gateway + staging + serverless pipeline'],
    ['MySQL', 'Managed', 'No-code / API', 'Database CDC ingestion', 'Gateway on classic compute'],
    ['PostgreSQL', 'Managed', 'No-code / API', 'Database CDC ingestion', 'AWS/Azure only'],
    ['...17 total managed', 'Managed', 'No-code / API', 'SaaS apps & databases', 'See Cloud Availability sheet'],
]

row = 4
for d in standard_data:
    for col, val in enumerate(d, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER
        if d[1] == 'Standard':
            cell.fill = PatternFill('solid', fgColor='E8F4FD')
        elif d[1] == 'Managed':
            cell.fill = PatternFill('solid', fgColor='FFF2E8')
    row += 1

# Comparison summary
row += 2
comparisons = [
    ['Dimension', 'Standard Connectors', 'Managed Connectors'],
    ['Setup', 'Code required (Python/Scala/SQL)', 'UI wizard or API config — no code'],
    ['Customization', 'Full control over transformations', 'Limited to connector options'],
    ['Compute', 'Your clusters or serverless', 'Serverless (pipeline) + classic (gateway for DB)'],
    ['Sources', 'Files, Kafka, Kinesis, Pub/Sub', 'SaaS apps + databases (17 connectors)'],
    ['Schema Evolution', 'Manual or Auto Loader inference', 'Automatic with managed rules'],
    ['Monitoring', 'Spark UI + custom', 'Built-in event logs + health metrics'],
    ['Cost Model', 'DBU based on cluster/serverless', 'DBU based on serverless pipeline runs'],
    ['Best For', 'Custom pipelines, streaming, files', 'Quick SaaS/DB ingestion with minimal code'],
]

for r, comp in enumerate(comparisons):
    for col, val in enumerate(comp, 1):
        cell = ws3.cell(row=row+r, column=col, value=val)
        cell.border = THIN_BORDER
        if r == 0:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        else:
            cell.font = BODY_FONT
            cell.alignment = LEFT

ws3.column_dimensions['A'].width = 38
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 40
ws3.column_dimensions['E'].width = 50

# ════════════════════════════════════════════════════
# Sheet 4: Interview Quick Reference
# ════════════════════════════════════════════════════
ws4 = wb.create_sheet('Interview Quick Ref')

ws4.merge_cells('A1:B1')
ws4['A1'].value = 'Lakeflow Connect — Interview Quick Reference'
ws4['A1'].font = Font(bold=True, size=14, name='Calibri', color='1B3A5C')

talking_points = [
    ['What is Lakeflow Connect?', 'Unified ingestion layer in Databricks. Two tiers: Standard connectors (Auto Loader, Kafka, Kinesis, Pub/Sub) for code-based pipelines, and Managed connectors (17 SaaS/DB sources) for zero-code ingestion.'],
    ['How many connectors?', '17 managed connectors across SaaS + databases. 5 GA (Salesforce, ServiceNow, Google Analytics, SQL Server, Workday). Plus 5 standard connectors (Auto Loader, Kafka, Kinesis, Pub/Sub, SFTP).'],
    ['Cloud parity?', 'AWS leads with all 17. Azure has 16 (same set). GCP has 14 — missing Jira, NetSuite, PostgreSQL.'],
    ['Architecture (SaaS)?', 'Connection (UC securable) → Ingestion Pipeline (serverless) → Streaming Tables (Delta). Scheduled via Workflows.'],
    ['Architecture (Database)?', 'Connection → Gateway (classic compute, continuous CDC capture) → Staging Volume → Ingestion Pipeline (serverless) → Streaming Tables.'],
    ['How does it relate to SDP?', 'Managed connectors output streaming tables — the same construct used in Spark Declarative Pipelines. The ingestion pipeline IS an SDP pipeline under the hood.'],
    ['How does it relate to Auto Loader?', 'Auto Loader is a standard connector for file-based sources. Managed connectors handle SaaS APIs and database CDC — complementary, not competing.'],
    ['FinServ angle?', 'Database connectors (SQL Server, MySQL, PostgreSQL) are critical for FinServ migrations. CDC capture with staging volumes means zero data loss. UC governance on connections means audit trail for regulators.'],
    ['vs Fivetran/Airbyte?', 'Fewer connectors today (~17 vs 300+), but native integration: UC governance, serverless compute, Delta Lake output, Workflows orchestration. No data leaving the platform. For FinServ, "data never leaves Databricks" is a compliance win.'],
    ['Key limitation?', 'Connector count is still growing. For niche SaaS sources, customers may still need Fivetran/Airbyte partner connectors. Schema evolution doesn\'t support data type changes on any connector.'],
]

for col, h in enumerate(['Topic', 'Talking Point'], 1):
    cell = ws4.cell(row=3, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for r, tp in enumerate(talking_points, 4):
    ws4.cell(row=r, column=1, value=tp[0]).font = BOLD_FONT
    ws4.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws4.cell(row=r, column=1).border = THIN_BORDER
    ws4.cell(row=r, column=2, value=tp[1]).font = BODY_FONT
    ws4.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws4.cell(row=r, column=2).border = THIN_BORDER
    ws4.row_dimensions[r].height = 50

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 100

# ── Save ──
output = '/Users/slysik/databricks/Lakeflow_Connect_Matrix.xlsx'
wb.save(output)
print(f'Saved to {output}')
